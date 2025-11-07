from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict

try:
    import win32com.client  # type: ignore
except ImportError:  # pragma: no cover - linux fallback
    win32com = None  # type: ignore

from openpyxl import Workbook

logger = logging.getLogger(__name__)


SUMMARY_ADDRESSES = {
    "Sheet3!B2": ("Sheet3", "B2"),
    "Sheet3!B3": ("Sheet3", "B3"),
    "Sheet3!B4": ("Sheet3", "B4"),
    "Sheet3!B5": ("Sheet3", "B5"),
    "Sheet3!B6": ("Sheet3", "B6"),
    "Sheet3!B7": ("Sheet3", "B7"),
    "Sheet3!B8": ("Sheet3", "B8"),
    "Sheet3!B9": ("Sheet3", "B9"),
    "Sheet3!B10": ("Sheet3", "B10"),
    "Sheet3!B11": ("Sheet3", "B11"),
    "Sheet3!B12": ("Sheet3", "B12"),
    "Sheet3!B13": ("Sheet3", "B13"),
    "Sheet1!B12": ("Sheet1", "B12"),
}

SUMMARY_CELL_MAP = {
    "Sheet3": {
        "A1": "Sell Price List",
    },
    "Summary": {
        "A1": "Summary",
    },
}


class CostingWorkbookWriter:
    def __init__(self, allow_xlsb: bool = False):
        self.allow_xlsb = allow_xlsb and win32com is not None

    def write(self, export: Dict[str, float], output_path: Path) -> Path:
        if self.allow_xlsb:
            return self._write_xlsb(export, output_path.with_suffix(".xlsb"))
        return self._write_xlsx(export, output_path.with_suffix(".xlsx"))

    def _write_xlsb(self, export: Dict[str, float], path: Path) -> Path:  # pragma: no cover - requires Windows
        excel = win32com.client.Dispatch("Excel.Application")  # type: ignore[attr-defined]
        wb = excel.Workbooks.Add()
        summary = wb.Sheets.Add()
        summary.Name = "Summary"
        sell = wb.Sheets.Add()
        sell.Name = "Sell Price List"
        for address, value in export.items():
            sheet_name, cell = address.split("!")
            sheet = wb.Sheets(sheet_name)
            sheet.Range(cell).Value = value
        wb.SaveAs(str(path), FileFormat=50)
        wb.Close()
        excel.Quit()
        return path

    def _write_xlsx(self, export: Dict[str, float], path: Path) -> Path:
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"

        sheets = {summary_sheet.title: summary_sheet}

        for sheet_name in {"Sell Price List"}:
            sheets[sheet_name] = wb.create_sheet(sheet_name)

        for sheet_name, cell_map in SUMMARY_CELL_MAP.items():
            ws = sheets.get(sheet_name)
            if ws is None:
                ws = wb.create_sheet(sheet_name)
                sheets[sheet_name] = ws
            for cell, value in cell_map.items():
                ws[cell] = value

        for address, value in export.items():
            sheet_name, cell = address.split("!")
            ws = sheets.get(sheet_name)
            if ws is None:
                ws = wb.create_sheet(sheet_name)
                sheets[sheet_name] = ws
            ws[cell] = value
        wb.save(path)
        logger.info("Saved costing workbook to %s", path)
        return path
